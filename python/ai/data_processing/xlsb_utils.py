import os
import sys
import csv


try:
    import hashlib
except Exception as ex:
    print('Need install hashlib')

try:
    from openpyxl import Workbook
    from pyxlsb import open_workbook
except Exception as err:
    print(err)
    print('Try use: source activate xlsbTools')



def mapXLSB(file, hash_columns, path=os.path.dirname(os.path.abspath(__file__)), save_path=None, print_interval=1000):
    if save_path is None:
        save_path = path

    with open_workbook(path+file) as wb:
        wbx = Workbook()

        for s in wb.sheets:
            header_index = {}
            hashed_columns = []
            columns_set = False
            wsx = wbx.create_sheet(s)
            print('Reading Sheet ', s)

            with wb.get_sheet(s) as sheet:
                row_count = 0
                # vrow = [0 for x in range(sheet.dimension.w)]
                sheet = list(sheet)
                for row in sheet:
                    # print(sheet.index(row))
                    if(row_count % print_interval == 0):
                        print('Reading row: ', row_count)
                    row_count += 1

                    if not columns_set:
                        for c in row:
                            wsx.cell(row=1+c.r, column=1+c.c).value = c.v
                        for i in range(0, len(row)):  # Map Columns By Name
                            header_index[row[i].v] = i
                        hashed_columns = sorted([  # Order and Select Hashed Column indexes
                            header_index[key] for key in header_index.keys()
                            if key in hash_columns])
                        columns_set = True
                    else:
                        for c in row:
                            str_value = str(c.v).encode()
                            if c.c in hashed_columns and s:
                                wsx.cell(row=1+c.r, column=1+c.c).value = hashlib.md5(
                                    str_value).hexdigest()
                            else:
                                try:
                                    wsx.cell(row=1+c.r, column=1 +
                                             c.c).value = c.v
                                except Exception as ex:
                                    print(ex)

        removeEmptySheet(wb, wbx)
        try:
            fname = save_path+file+'.xls'
            print('... Saving: ', fname)
            wbx.save(fname)
            print('Sucess!')
        except Exception as ex:
            print(ex)
            print('Saving Failed')


def removeEmptySheet(wb, wbx):
    for sheetName in filter(lambda n: n not in wb.sheets, wbx.get_sheet_names()):
        try:
            s = wbx.get_sheet_by_name(sheetName)
            wbx.remove_sheet(s)
        except Exception as ex:
            print(ex)
            print(sheetName)


def xlsbToXlsx(path_in, path_out,separate_sheets=False, print_interval=1000):
    print('parsing: ')
    print(path_in)
    print(path_out)
    if separate_sheets :
        print('Saving in differents sheets')
    
    with open_workbook(path_in) as wb:
        wbx = Workbook()

        for s in wb.sheets:
            wsx = wbx.create_sheet(s)
            print('Reading Sheet ', s)

            with wb.get_sheet(s) as sheet:
                row_count = 0
                # vrow = [0 for x in range(sheet.dimension.w)]
                sheet = list(sheet)
                for row in sheet:
                    # print(sheet.index(row))
                    if(row_count % print_interval == 0):
                        print('Reading row: ', row_count)
                    row_count += 1
                    for c in row:
                            wsx.cell(row=1+c.r, column=1+c.c).value = c.v
            if (separate_sheets):
                print('Saving partials...')
                removeEmptySheet(wb, wbx)
                wbx.save(path_out+"."+s+".xls")
                wbx = Workbook()
                print('Saving Completed.')
        try:
            if (~separate_sheets):
                print('... Saving: ', path_out)
                removeEmptySheet(wb, wbx)
                wbx.save(path_out+'.xls')
                
                print('Sucess!')
        except Exception as ex:
            print(ex)
            print('Saving Failed')


def xlsbToCSV(path_in, path_out,sep=";", print_interval=1000):
    print('parsing to csv..')
    row_count = 0
    with open_workbook(path_in) as wb:
        for s in wb.sheets:
            with wb.get_sheet(s) as sheet:
                print('reading sheet',s)
                with open(path_out+'/'+s+'.csv', 'w', encoding='utf-8') as csvfile:
                    spamwriter = csv.writer(csvfile, delimiter=sep, quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    sheet = list(sheet)
                    for row in sheet:
                        if(row_count % print_interval == 0):
                            print('Reading row: ', row_count)
                        row_count += 1
                        spamwriter.writerow([c.v for c in row])


                        
                

import sys
if __name__ == '__main__':
    
    # xls_utils.py 0
    # FUNCTIONNAME 1 
    # print(sys.argv)
    # parse_xlsb
    
    print(sys.argv)
    if( sys.argv[1] == 'xls'): 
        from_path = sys.argv[2]
        to_path = sys.argv[3]
        xlsbToXlsx(from_path,to_path,separate_sheets=True,print_interval=500)
    elif( sys.argv[1] == 'csv'): 
        from_path = sys.argv[2]
        to_path = sys.argv[3]
        xlsbToCSV(from_path,to_path,print_interval=500)
